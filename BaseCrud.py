from datetime import datetime
from openpyxl import load_workbook
rut=r'DatosCrud.xlsx'

def leer(ruta:str, extraer:str):
 Archivo_Excel=load_workbook(ruta)
 Hoja_datos=Archivo_Excel['Datos del crud']
 Hoja_datos=Hoja_datos['A2' : 'F'+str(Hoja_datos.max_row)]

 info={}

 for i in Hoja_datos:

     if isinstance(i[0].value,int):
        info.setdefault(i[0].value,{'tarea':i[1].value, 'descripcion':i[2].value,
                                    'estado':i[3].value, 'fecha':i[4].value, 'fecha de finalizacion':i[5].value})
 if not(extraer=='todo'):
    info=filtrar(info,extraer)

 for i in info:
    print('******** Tarea ********')
    print('id:' + str(i)+ '\n' + 'Titulo: '+str(info[i]['tarea'])+ '\n'+'Descripcion: '+str(info[i]['descripcion'])+ '\n'+ 'Estado: '+str(info[i]['estado'])
           +'\n' +'Fecha Creacion: '+str(info[i]['fecha'])
           +'\n' +'Fecha de finalizacion: '+str(info [i]['fecha de finalizacion']))
    print()
 return 
def filtrar(info:dict,filtro:str):
    aux={}
    for i in info:
        if info[i]['estado']==filtro:
          aux.setdefault(i,info[i])
    return aux

def actualizar(ruta:str,identificador:int,datos_actualizados:dict):
  Archivo_Excel=load_workbook(ruta)
  Hoja_datos=Archivo_Excel['Datos del crud']
  Hoja_datos=Hoja_datos['A2' : 'F'+str(Hoja_datos.max_row)]
  hoja=Archivo_Excel.active3

  titulo=2
  descripcion=3
  estado=4
  fecha_inicio=5
  fecha_de_finalizacion=6
  encontro=False
  for i in Hoja_datos:
      if i[0].value==identificador:
          fila=i[0].row
          encontro=True
          for d in datos_actualizados:
              if d=='titulo' and not(datos_actualizados[d]==''):
                hoja.cell(row=fila,column=titulo).value=datos_actualizados[d]
              elif d=='descripcion' and not(datos_actualizados[d]==''):
                hoja.cell(row=fila,column=descripcion).value=datos_actualizados[d]
              elif d=='estado' and not(datos_actualizados[d]==''):
                hoja.cell(row=fila,column=estado).value=datos_actualizados[d]
              elif d=='fecha_inicio' and not(datos_actualizados[d]==''):
                hoja.cell(row=fila,column=fecha_inicio).value=datos_actualizados[d]
              elif d=='fecha_de_finalizacion' and not(datos_actualizados[d]==''):
                hoja.cell(row=fila,column=fecha_de_finalizacion).value=datos_actualizados[d]
  Archivo_Excel.save(ruta)
  if encontro==False:
    print('Error: No existe una tarea con ese Id')
    print()
  return

def agregar(ruta:int, datos:dict):
  Archivo_Excel=load_workbook(ruta)
  Hoja_datos=Archivo_Excel['Datos del crud']
  Hoja_datos=Hoja_datos['A2' : 'F'+str(Hoja_datos.max_row+1)]
  hoja=Archivo_Excel.active

  titulo=2
  descripcion=3
  estado=4
  fecha_inicio=5
  fecha_de_finalizacion=6
  for i in Hoja_datos:

      if not(isinstance(i[0].value,int)):
         identificador=i[0].row
         hoja.cell(row=identificador,column=1).value=identificador-1
         hoja.cell(row=identificador,column=titulo).value=datos['titulo']
         hoja.cell(row=identificador,column=descripcion).value=datos['descripcion']
         hoja.cell(row=identificador,column=estado).value=datos['estado']
         hoja.cell(row=identificador,column=fecha_inicio).value=datos['fecha inicio']
         hoja.cell(row=identificador,column=fecha_de_finalizacion).value=datos['fecha finalizado']
         break
  Archivo_Excel.save(ruta)
  return

def borrar(ruta,identificador):
  Archivo_Excel=load_workbook(ruta)
  Hoja_datos=Archivo_Excel['Datos del crud']
  Hoja_datos=Hoja_datos['A2' : 'F'+str(Hoja_datos.max_row)]
  hoja=Archivo_Excel.active

  titulo=2
  descripcion=3
  estado=4
  fecha_inicio=5
  fecha_de_finalizacion=6
  encontro=False
  for i in Hoja_datos:
      if i [0].value==identificador:
          fila=i[0].row
          encontro=True

          hoja.cell(row=fila, column=1).value=''
          hoja.cell(row=fila, column=titulo).value=''
          hoja.cell(row=fila, column=descripcion).value=' ' 
          hoja.cell(row=fila, column=estado).value=''
          hoja.cell(row=fila, column=fecha_inicio).value=''
          hoja.cell(row=fila, column=fecha_de_finalizacion).value=''
  Archivo_Excel.save(ruta)
  if encontro==False:
    print('Error: No existe una tarea con ese Id')
    print()
  return



datos_actualizados={'titulo':'', 'descripcion': '', 'estado': '', 'fecha_inicio': '', 'fecha_finalizacion': ''}
while True:
  print('Indique la accion que desea realizar')
  print('Consultar: 1')
  print('Actualizar: 2')
  print('crear nueva tarea: 3')
  print('Borrar: 4')
  accion=input('Escriba la accion:  ')
  if not(accion=='1') and not(accion=='2') and not(accion=='3') and not(accion=='4'):
    print('comando invalida por favor elija una opcion valida')
  elif accion=='1':
         opc_consulta=''
         print('Indique la tarea que desea consultar: ')
         print('Todas las tareas: 1')
         print('En espera: 2')
         print('En ejecucion: 3')
         print('Por aprobar: 4')
         print('Finalizada: 5')
         opc_consulta = input('Escriba la tarea que dese consultar: ')
         if opc_consulta=='1':
          print()
          print()
          print('** Consultado todaslas tareas **')
          leer(rut,'todo')

         elif opc_consulta=='2':
          print()
          print()
          print('** Consultado todaslas tareas **')
          leer(rut,'En espera')

         elif opc_consulta=='3':
          print()
          print()
          print('** Consultado todaslas tareas **')
          leer(rut,'En ejecucion')

         elif opc_consulta=='4':
          print()
          print()
          print('** Consultado todaslas tareas **')
          leer(rut,'Poir aprobar')

         elif opc_consulta=='5':
          print()
          print()
          print('** Consultado todaslas tareas **')
          leer(rut,'Finalizada')
  elif accion=='2':
     datos_Actualizados={'titulo':'', 'descripcion':'', 'estado':'', 'fecha inicio':'', 'fecha finalizacion':''}
     print('** Actualizar tarea **')
     print()
     id_Actualizar=int(input( 'Indique el id de la tarea que desea actualizar: '))
     print()
     print('** Nuevo titulo **')
     print('**Nota: si no desea actualizar el titulo solo oprima ENTER')
     datos_Actualizados['descripcion']= input('Indique la nueva descripcion de la tarea : ')
     print()
     print()
     print('** Nuevo estado **')
     print('En espera: 2')
     print('En ejecucion: 3')
     print('Por aprobar: 4')
     print('Finalizada: 5')
     print('**Nota: si no desea actualizar el estado solo oprima ENTER')
     estadoNuevo=input('Indique el nuevo estado de la tarea: ')
     if estadoNuevo=='2':
      datos_Actualizados['estado']='En espera'
     elif estadoNuevo=='3':
      datos_Actualizados['estado']='En espera'
     elif estadoNuevo=='4':
      datos_Actualizados['estado']='En espera'
     elif estadoNuevo=='5':
      datos_Actualizados['estado']='En espera'
     now = datetime.now()
     datos_Actualizados['estado']='Finalizada'
     datos_Actualizados['fecha finalizacion']=str(now.day) +'/'+ str(now.month)+'/'+str(now.year)
     now = datetime.now()
     datos_Actualizados['fecha inicio']=str(now.day) +'/'+ str(now.month)+'/'+str(now.year)
     actualizar(rut,id_Actualizar, datos_Actualizados)
     print()

  elif accion=='3':
       datos_Actualizados={'titulo':'', 'descripcion':'', 'estado':'', 'fecha inicio':'', 'fecha finalizacion':''}
       print('** Crear nueva Tarea **')

       print()
       print('** titulo **')
       print()
       datos_Actualizados['titulo']=input('Indique el titulo de la tarea: ')
       print()
       print('** descripcion **')
       datos_Actualizados['descripcion']=input('Indique la descripcion de la tarea: ')
       print()
       datos_Actualizados['estado']='En espera'
       now = datetime.now()
       datos_Actualizados['fecha inicio']=str(now.day)+'/'+ str(now.month)+'/'+str(now.year)
       datos_Actualizados['fecha finalizacion']=''
       agregar(rut,datos_Actualizados)

  elif accion=='4':
      print('')
      print('** Eliminar Tarea **')
      iden=int(input('Indeique el Id de la tarea que desea eliminar'))
      borrar(rut,iden)




